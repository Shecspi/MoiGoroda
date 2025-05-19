import csv
import json
from datetime import datetime
from io import BytesIO

import openpyxl
import pytest
from django.urls import reverse

from account.tests.download.create_db import (
    create_user,
    create_area,
    create_region,
    create_city,
    create_visited_city,
)


@pytest.fixture
def setup_db(django_user_model):
    user = create_user(django_user_model, 1)
    area = create_area(1)
    region = create_region(1, area[0])
    city = create_city(2, region[0])
    date_of_visit_1 = datetime.strptime('2024-01-01', '%Y-%m-%d')
    date_of_visit_2 = datetime.strptime('2023-01-01', '%Y-%m-%d')
    create_visited_city(region[0], user, city[0], date_of_visit_1, False, 3)
    create_visited_city(region[0], user, city[1], date_of_visit_2, True, 5)


@pytest.mark.django_db
def test__content_of_city_report_in_txt_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'city', 'filetype': 'txt'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode().split('\n')

    assert (
        report[0]
        == 'Город       Регион               Дата посещения     Наличие сувенира     Оценка     '
    )
    assert (
        report[1]
        == 'Город 1     Регион 1 область     2024-01-01         -                    ***        '
    )
    assert (
        report[2]
        == 'Город 2     Регион 1 область     2023-01-01         +                    *****      '
    )


@pytest.mark.django_db
def test__content_of_city_report_in_csv_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'city', 'filetype': 'csv'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode().strip().split('\n')
    reader = csv.reader(report, delimiter=',', lineterminator='\n')
    expected_results = (
        ['Город', 'Регион', 'Дата посещения', 'Наличие сувенира', 'Оценка'],
        ['Город 1', 'Регион 1 область', '2024-01-01', '-', '***'],
        ['Город 2', 'Регион 1 область', '2023-01-01', '+', '*****'],
    )

    for index, row in enumerate(reader):
        assert row == expected_results[index]


@pytest.mark.django_db
def test__content_of_city_report_in_json_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'city', 'filetype': 'json'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode()
    recieved_data = json.loads(report)
    expected_data = [
        ['Город', 'Регион', 'Дата посещения', 'Наличие сувенира', 'Оценка'],
        ['Город 1', 'Регион 1 область', '2024-01-01', '-', '***'],
        ['Город 2', 'Регион 1 область', '2023-01-01', '+', '*****'],
    ]

    assert recieved_data == expected_data


@pytest.mark.django_db
def test__content_of_city_report_in_xls_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'city', 'filetype': 'xls'}
    response = client.post(reverse('download'), data=data)
    workbook = openpyxl.load_workbook(BytesIO(response.content))
    expected_data = [
        ('Город', 'Регион', 'Дата посещения', 'Наличие сувенира', 'Оценка'),
        ('Город 1', 'Регион 1 область', '2024-01-01', '-', '***'),
        ('Город 2', 'Регион 1 область', '2023-01-01', '+', '*****'),
    ]

    assert list(workbook['Sheet'].values) == expected_data


@pytest.mark.django_db
def test__content_of_region_report_in_txt_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'region', 'filetype': 'txt'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode().split('\n')

    assert report[0] == (
        'Регион               Всего городов     Посещено городов, шт     '
        'Посещено городов, %     Осталось посетить, шт     '
    )
    assert report[1] == (
        'Регион 1 область     2                 2                        '
        '100%                    0                         '
    )


@pytest.mark.django_db
def test__content_of_region_report_in_csv_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'region', 'filetype': 'csv'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode().strip().split('\n')
    reader = csv.reader(report, delimiter=',', lineterminator='\n')
    expected_results = (
        [
            'Регион',
            'Всего городов',
            'Посещено городов, шт',
            'Посещено городов, %',
            'Осталось посетить, шт',
        ],
        ['Регион 1 область', '2', '2', '100%', '0'],
    )

    for index, row in enumerate(reader):
        assert row == expected_results[index]


@pytest.mark.django_db
def test__content_of_region_report_in_json_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'region', 'filetype': 'json'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode()
    recieved_data = json.loads(report)
    expected_data = [
        [
            'Регион',
            'Всего городов',
            'Посещено городов, шт',
            'Посещено городов, %',
            'Осталось посетить, шт',
        ],
        ['Регион 1 область', '2', '2', '100%', '0'],
    ]

    assert recieved_data == expected_data


@pytest.mark.django_db
def test__content_of_region_report_in_xls_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'region', 'filetype': 'xls'}
    response = client.post(reverse('download'), data=data)
    workbook = openpyxl.load_workbook(BytesIO(response.content))
    expected_data = [
        (
            'Регион',
            'Всего городов',
            'Посещено городов, шт',
            'Посещено городов, %',
            'Осталось посетить, шт',
        ),
        ('Регион 1 область', '2', '2', '100%', '0'),
    ]

    assert list(workbook['Sheet'].values) == expected_data


@pytest.mark.django_db
def test__content_of_area_report_in_txt_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'area', 'filetype': 'txt'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode().split('\n')

    assert report[0] == (
        'Федеральный округ     Всего регионов, шт     Посещено регионов, шт     '
        'Посещено регионов, %     Осталось посетить, шт     '
    )
    assert report[1] == (
        'Округ 1               1                      1                         '
        '100%                     0                         '
    )


@pytest.mark.django_db
def test__content_of_area_report_in_csv_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'area', 'filetype': 'csv'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode().strip().split('\n')
    reader = csv.reader(report, delimiter=',', lineterminator='\n')
    expected_results = (
        [
            'Федеральный округ',
            'Всего регионов, шт',
            'Посещено регионов, шт',
            'Посещено регионов, %',
            'Осталось посетить, шт',
        ],
        ['Округ 1', '1', '1', '100%', '0'],
    )

    for index, row in enumerate(reader):
        assert row == expected_results[index]


@pytest.mark.django_db
def test__content_of_area_report_in_json_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'area', 'filetype': 'json'}
    response = client.post(reverse('download'), data=data)
    report = response.content.decode()
    recieved_data = json.loads(report)
    expected_data = [
        [
            'Федеральный округ',
            'Всего регионов, шт',
            'Посещено регионов, шт',
            'Посещено регионов, %',
            'Осталось посетить, шт',
        ],
        ['Округ 1', '1', '1', '100%', '0'],
    ]

    assert recieved_data == expected_data


@pytest.mark.django_db
def test__content_of_area_report_in_xls_file(setup_db, client):
    client.login(username='username1', password='password')
    data = {'reporttype': 'area', 'filetype': 'xls'}
    response = client.post(reverse('download'), data=data)
    workbook = openpyxl.load_workbook(BytesIO(response.content))
    expected_data = [
        (
            'Федеральный округ',
            'Всего регионов, шт',
            'Посещено регионов, шт',
            'Посещено регионов, %',
            'Осталось посетить, шт',
        ),
        ('Округ 1', '1', '1', '100%', '0'),
    ]

    assert list(workbook['Sheet'].values) == expected_data
