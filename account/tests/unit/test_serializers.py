"""
----------------------------------------------

Copyright © Egor Vavilov (Shecspi)
Licensed under the Apache License, Version 2.0

----------------------------------------------
"""

import openpyxl
import pytest
import json
from io import StringIO, BytesIO

from account.serializer import TxtSerializer, CsvSerializer, XlsSerializer, JsonSerializer


# ===== Фикстуры =====


@pytest.fixture
def sample_report() -> list[tuple[str, ...]]:
    """Общая фикстура с тестовыми данными для отчётов"""
    return [
        ('Город', 'Регион', 'Количество посещений'),
        ('Москва', 'Москва', '5'),
        ('Жуков', 'Калужская область', '10'),
    ]


@pytest.fixture
def empty_report() -> list[tuple[str, ...]]:
    """Фикстура с пустым отчётом (только заголовки)"""
    return [('Город', 'Регион', 'Количество посещений')]


@pytest.fixture
def complex_report() -> list[tuple[str, ...]]:
    """Фикстура с более сложными данными"""
    return [
        ('Город', 'Регион', 'Количество', 'Процент', 'Дополнительно'),
        ('Москва', 'Москва', '100', '50%', 'Столица'),
        ('Санкт-Петербург', 'Ленинградская область', '50', '25%', 'Северная столица'),
        ('Казань', 'Республика Татарстан', '25', '12.5%', 'Город спорта'),
    ]


# ===== Тесты для TxtSerializer =====


@pytest.mark.unit
def test_txt_serializer_content_type():
    """Тест метода content_type для TxtSerializer"""
    serializer = TxtSerializer()
    assert serializer.content_type() == 'text/txt'


@pytest.mark.unit
def test_txt_serializer_filetype():
    """Тест метода filetype для TxtSerializer"""
    serializer = TxtSerializer()
    assert serializer.filetype() == 'txt'


@pytest.mark.unit
def test_txt_serializer_convert(sample_report):
    """Тест метода convert для TxtSerializer"""
    serializer = TxtSerializer()
    result = serializer.convert(sample_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()

    expected_content = (
        'Город      Регион                Количество посещений     \n'
        'Москва     Москва                5                        \n'
        'Жуков      Калужская область     10                       \n'
    )
    assert content == expected_content


@pytest.mark.unit
def test_txt_serializer_get_max_length(sample_report):
    """Тест приватного метода __get_max_length для TxtSerializer"""
    serializer = TxtSerializer()
    result = serializer._TxtSerializer__get_max_length(sample_report)
    assert result == [6, 17, 20]


@pytest.mark.unit
def test_txt_serializer_get_formatted_row():
    """Тест приватного метода __get_formated_row для TxtSerializer"""
    serializer = TxtSerializer()
    formatted_row = serializer._TxtSerializer__get_formated_row(
        ('Жуков', 'Калужская область', '10'), [6, 17, 20]
    )
    assert formatted_row == 'Жуков      Калужская область     10                       \n'


@pytest.mark.unit
def test_txt_serializer_empty_report(empty_report):
    """Тест TxtSerializer с пустым отчётом"""
    serializer = TxtSerializer()
    result = serializer.convert(empty_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()
    assert 'Город' in content
    assert 'Регион' in content


@pytest.mark.unit
def test_txt_serializer_complex_report(complex_report):
    """Тест TxtSerializer со сложным отчётом"""
    serializer = TxtSerializer()
    result = serializer.convert(complex_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()

    assert 'Москва' in content
    assert 'Санкт-Петербург' in content
    assert 'Казань' in content


# ===== Тесты для CsvSerializer =====


@pytest.mark.unit
def test_csv_serializer_content_type():
    """Тест метода content_type для CsvSerializer"""
    serializer = CsvSerializer()
    assert serializer.content_type() == 'text/csv'


@pytest.mark.unit
def test_csv_serializer_filetype():
    """Тест метода filetype для CsvSerializer"""
    serializer = CsvSerializer()
    assert serializer.filetype() == 'csv'


@pytest.mark.unit
def test_csv_serializer_convert(sample_report):
    """Тест метода convert для CsvSerializer"""
    serializer = CsvSerializer()
    result = serializer.convert(sample_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()

    expected_content = (
        'Город,Регион,Количество посещений\nМосква,Москва,5\nЖуков,Калужская область,10\n'
    )
    assert content == expected_content


@pytest.mark.unit
def test_csv_serializer_empty_report(empty_report):
    """Тест CsvSerializer с пустым отчётом"""
    serializer = CsvSerializer()
    result = serializer.convert(empty_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()
    assert 'Город,Регион,Количество посещений\n' == content


@pytest.mark.unit
def test_csv_serializer_complex_report(complex_report):
    """Тест CsvSerializer со сложным отчётом"""
    serializer = CsvSerializer()
    result = serializer.convert(complex_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()

    lines = content.split('\n')
    assert len(lines) == 5  # 4 строки данных + пустая строка в конце
    assert 'Санкт-Петербург' in content


@pytest.mark.unit
def test_csv_serializer_special_characters():
    """Тест CsvSerializer с специальными символами"""
    serializer = CsvSerializer()
    report = [
        ('Город', 'Описание'),
        ('Москва', 'Город, который является столицей'),
    ]
    result = serializer.convert(report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()
    # CSV должен правильно обрабатывать запятые
    assert 'Москва' in content


# ===== Тесты для XlsSerializer =====


@pytest.mark.unit
def test_xls_serializer_content_type():
    """Тест метода content_type для XlsSerializer"""
    serializer = XlsSerializer()
    assert serializer.content_type() == 'application/vnd.ms-excel'


@pytest.mark.unit
def test_xls_serializer_filetype():
    """Тест метода filetype для XlsSerializer"""
    serializer = XlsSerializer()
    assert serializer.filetype() == 'xls'


@pytest.mark.unit
def test_xls_serializer_convert(sample_report):
    """Тест метода convert для XlsSerializer"""
    serializer = XlsSerializer()
    result = serializer.convert(sample_report)

    assert isinstance(result, BytesIO)
    result.seek(0)

    # Открываем созданный файл как Excel
    workbook = openpyxl.load_workbook(result)
    worksheet = workbook.active

    # Проверяем заголовки
    assert worksheet['A1'].value == 'Город'
    assert worksheet['B1'].value == 'Регион'
    assert worksheet['C1'].value == 'Количество посещений'

    # Проверяем данные
    assert worksheet['A2'].value == 'Москва'
    assert worksheet['B2'].value == 'Москва'
    assert worksheet['C2'].value == '5'

    assert worksheet['A3'].value == 'Жуков'
    assert worksheet['B3'].value == 'Калужская область'
    assert worksheet['C3'].value == '10'


@pytest.mark.unit
def test_xls_serializer_empty_report(empty_report):
    """Тест XlsSerializer с пустым отчётом"""
    serializer = XlsSerializer()
    result = serializer.convert(empty_report)

    assert isinstance(result, BytesIO)
    result.seek(0)

    workbook = openpyxl.load_workbook(result)
    worksheet = workbook.active

    # Проверяем, что есть заголовки
    assert worksheet['A1'].value == 'Город'
    assert worksheet['B1'].value == 'Регион'
    assert worksheet['C1'].value == 'Количество посещений'

    # Вторая строка должна быть пустой
    assert worksheet['A2'].value is None


@pytest.mark.unit
def test_xls_serializer_complex_report(complex_report):
    """Тест XlsSerializer со сложным отчётом"""
    serializer = XlsSerializer()
    result = serializer.convert(complex_report)

    assert isinstance(result, BytesIO)
    result.seek(0)

    workbook = openpyxl.load_workbook(result)
    worksheet = workbook.active

    # Проверяем количество строк
    assert worksheet['A2'].value == 'Москва'
    assert worksheet['A3'].value == 'Санкт-Петербург'
    assert worksheet['A4'].value == 'Казань'


# ===== Тесты для JsonSerializer =====


@pytest.mark.unit
def test_json_serializer_content_type():
    """Тест метода content_type для JsonSerializer"""
    serializer = JsonSerializer()
    assert serializer.content_type() == 'application/json'


@pytest.mark.unit
def test_json_serializer_filetype():
    """Тест метода filetype для JsonSerializer"""
    serializer = JsonSerializer()
    assert serializer.filetype() == 'json'


@pytest.mark.unit
def test_json_serializer_convert(sample_report):
    """Тест метода convert для JsonSerializer"""
    serializer = JsonSerializer()
    result = serializer.convert(sample_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()

    expected_content = """[
    [
        "Город",
        "Регион",
        "Количество посещений"
    ],
    [
        "Москва",
        "Москва",
        "5"
    ],
    [
        "Жуков",
        "Калужская область",
        "10"
    ]
]"""
    assert content == expected_content


@pytest.mark.unit
def test_json_serializer_valid_json(sample_report):
    """Тест что JsonSerializer генерирует валидный JSON"""
    serializer = JsonSerializer()
    result = serializer.convert(sample_report)

    result.seek(0)
    content = result.read()

    # Проверяем, что это валидный JSON
    parsed = json.loads(content)
    assert isinstance(parsed, list)
    assert len(parsed) == 3


@pytest.mark.unit
def test_json_serializer_empty_report(empty_report):
    """Тест JsonSerializer с пустым отчётом"""
    serializer = JsonSerializer()
    result = serializer.convert(empty_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()

    parsed = json.loads(content)
    assert len(parsed) == 1
    assert parsed[0] == ['Город', 'Регион', 'Количество посещений']


@pytest.mark.unit
def test_json_serializer_complex_report(complex_report):
    """Тест JsonSerializer со сложным отчётом"""
    serializer = JsonSerializer()
    result = serializer.convert(complex_report)

    assert isinstance(result, StringIO)
    result.seek(0)
    content = result.read()

    parsed = json.loads(content)
    assert len(parsed) == 4
    assert parsed[1][0] == 'Москва'
    assert parsed[2][0] == 'Санкт-Петербург'


@pytest.mark.unit
def test_json_serializer_unicode():
    """Тест JsonSerializer с unicode символами"""
    serializer = JsonSerializer()
    report = [
        ('Город', 'Описание'),
        ('Москва', 'Столица России 🇷🇺'),
    ]
    result = serializer.convert(report)

    result.seek(0)
    content = result.read()

    parsed = json.loads(content)
    assert 'Столица России 🇷🇺' in parsed[1]


# ===== Общие тесты для всех сериализаторов =====


@pytest.mark.unit
def test_all_serializers_implement_interface():
    """Тест что все сериализаторы реализуют необходимые методы"""
    serializers = [TxtSerializer(), CsvSerializer(), XlsSerializer(), JsonSerializer()]

    for serializer in serializers:
        assert hasattr(serializer, 'convert')
        assert hasattr(serializer, 'content_type')
        assert hasattr(serializer, 'filetype')
        assert callable(serializer.convert)
        assert callable(serializer.content_type)
        assert callable(serializer.filetype)


@pytest.mark.unit
def test_all_serializers_return_correct_types(sample_report):
    """Тест что все сериализаторы возвращают правильные типы"""
    txt_serializer = TxtSerializer()
    csv_serializer = CsvSerializer()
    json_serializer = JsonSerializer()
    xls_serializer = XlsSerializer()

    # Текстовые сериализаторы возвращают StringIO
    assert isinstance(txt_serializer.convert(sample_report), StringIO)
    assert isinstance(csv_serializer.convert(sample_report), StringIO)
    assert isinstance(json_serializer.convert(sample_report), StringIO)

    # XLS сериализатор возвращает BytesIO
    assert isinstance(xls_serializer.convert(sample_report), BytesIO)
